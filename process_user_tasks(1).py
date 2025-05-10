import os
import re
import json
import pandas as pd
import glob
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font


def extract_data_from_file(file_path):
    """
    Extract required data from a txt file which contains JSON conversation history
    Extract all dialog pairs from the conversation
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            
        # Skip empty files
        if not content or content == "[]":
            print(f"Skipping empty file: {file_path}")
            return None
            
        # Try to parse as JSON
        try:
            data = json.loads(content)
            
            # Extract case_id from filename
            case_id = os.path.basename(file_path).split('_')[0]
            
            # Based on the structure provided, we'll extract task number which maps to week
            task_num = extract_task_number(file_path)
            week = task_num  # task1 = week1, task2 = week2, etc.
            
            # Default timestamp, will be adjusted later
            timestamp = "01.01.25 09:00"
            
            # Extract all user-assistant pairs from the conversation
            conversation_pairs = []
            
            for i, message in enumerate(data):
                if message["role"] == "user" and i+1 < len(data):
                    user_msg = message["content"]
                    chatgpt_after = ""
                    chatgpt_before = ""
                    
                    # Get chatgpt_after (response after user message)
                    if i+1 < len(data) and data[i+1]["role"] == "assistant":
                        chatgpt_after = data[i+1]["content"]
                    
                    # Get chatgpt_before (response before user message)
                    if i-1 >= 0 and data[i-1]["role"] == "assistant":
                        chatgpt_before = data[i-1]["content"]
                    elif i-1 >= 0 and data[i-1]["role"] == "system":
                        chatgpt_before = data[i-1]["content"]
                    
                    # Add this pair to the results
                    conversation_pairs.append({
                        'case_id': case_id,
                        'week': str(week),
                        'idx': str(i),  # Use message index as temporary idx
                        'timestamp': timestamp,  # Will be adjusted later
                        'chatgpt_before': chatgpt_before,
                        'user': user_msg,
                        'chatgpt_after': chatgpt_after
                    })
            
            return conversation_pairs
            
        except json.JSONDecodeError:
            print(f"File {file_path} is not valid JSON. Trying alternate parsing...")
            
            # If not valid JSON, try the alternative approach (regular expression pattern)
            pattern = r"case_id\s+(.*?)\s+week\s+(.*?)\s+idx\s+(.*?)\s+timestamp\s+(.*?)\s+chatgpt_before\s+(.*?)\s+user\s+(.*?)\s+chatgpt_after\s+(.*)"
            match = re.search(pattern, content, re.DOTALL)
            
            if match:
                return [{
                    'case_id': match.group(1).strip(),
                    'week': match.group(2).strip(),
                    'idx': match.group(3).strip(),
                    'timestamp': match.group(4).strip(),
                    'chatgpt_before': match.group(5).strip(),
                    'user': match.group(6).strip(),
                    'chatgpt_after': match.group(7).strip()
                }]
            else:
                print(f"Could not parse file {file_path} with either method.")
                return None
                
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return None


def extract_user_id_from_filename(filename):
    """
    Extract the user ID from the filename.
    Assumes filename format like: 31de0a7aaa_task1.txt
    """
    # Split the filename by underscore and take the first part
    base_name = os.path.basename(filename)
    parts = base_name.split('_')
    if len(parts) > 0:
        return parts[0]
    return None


def extract_task_number(filename):
    """
    Extract the task number from the filename.
    Assumes filename format like: 31de0a7aaa_task1.txt or path/to/task1/31de0a7aaa.txt
    """
    # Check if task number is in the filename
    match = re.search(r'task(\d+)', filename)
    if match:
        return int(match.group(1))
    
    # Check if task number is in the directory path
    path = Path(filename)
    for parent in path.parents:
        match = re.search(r'task(\d+)', str(parent))
        if match:
            return int(match.group(1))
    
    return 1  # Default to task1 if not found


def process_all_files(base_dir):
    """
    Process all txt files in the task directories
    """
    all_data = []
    
    # Find the main directory
    main_dir = os.path.join(base_dir, "Student_GPT_conversation_by_question")
    if os.path.isdir(main_dir):
        base_dir = main_dir
    
    # Find task directories
    task_dirs = []
    for i in range(1, 19):  # task1 to task18
        task_dir = os.path.join(base_dir, f"task{i}")
        if os.path.isdir(task_dir):
            task_dirs.append((task_dir, i))
    
    # Process files in task directories
    for task_dir, task_num in task_dirs:
        print(f"Processing files in {task_dir}...")
        txt_files = glob.glob(os.path.join(task_dir, "*.txt"))
        for file_path in txt_files:
            case_id = extract_user_id_from_filename(file_path)
            if case_id:
                conversation_pairs = extract_data_from_file(file_path)
                if conversation_pairs:  # Only add if data is not None (skips empty files)
                    # Add each pair to the result list
                    for pair in conversation_pairs:
                        pair['case_id'] = case_id
                        pair['week'] = str(task_num)
                        all_data.append(pair)
    
    return all_data


def adjust_timestamps(data_list):
    """
    Adjust timestamps based on week and idx
    Inside same week for same user, increment by 10 minutes for each conversation
    """
    # Group data by case_id and week
    data_by_user_week = {}
    for item in data_list:
        case_id = item.get('case_id', '')
        week = int(item.get('week', 1))
        key = f"{case_id}_{week}"
        
        if key not in data_by_user_week:
            data_by_user_week[key] = []
        
        data_by_user_week[key].append(item)
    
    # Process each group
    adjusted_data = []
    for key, items in data_by_user_week.items():
        case_id, week_str = key.split('_')
        week = int(week_str)
        
        # Sort items by the temporary idx within the file
        items.sort(key=lambda x: int(x.get('idx', 0)))
        
        # Get base timestamp for this week
        base_timestamp_str = '01.01.25 09:00'  # Default start time
        try:
            # Parse the timestamp
            base_timestamp = datetime.strptime(base_timestamp_str, '%d.%m.%y %H:%M')
            
            # Adjust base timestamp for week (add 7 days for week 2, 14 days for week 3, etc.)
            if week > 1:
                days_to_add = (week - 1) * 7
                base_timestamp += timedelta(days=days_to_add)
            
            # Adjust timestamps for each conversation in this week
            for i, item in enumerate(items):
                # Increment by 10 minutes for each conversation in the same week
                adjusted_timestamp = base_timestamp + timedelta(minutes=i * 10)
                
                # Update the timestamp
                item['timestamp'] = adjusted_timestamp.strftime('%d.%m.%y %H:%M')
                
                # Add to adjusted data
                adjusted_data.append(item)
                
        except Exception as e:
            print(f"Error adjusting timestamps for {key}: {e}")
            # If error occurs, still add the items to the result but without adjusted timestamps
            for item in items:
                adjusted_data.append(item)
    
    return adjusted_data


def save_to_excel(data, output_file):
    """
    Save the processed data to an Excel file with properly formatted columns
    and truly automatic row heights
    """
    if not data:
        print("No data to save")
        return
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Sort by case_id and then by week
    if 'case_id' in df.columns and 'week' in df.columns:
        df['week'] = df['week'].astype(int)
        df = df.sort_values(by=['case_id', 'week'])
    
    # Reset idx to be simple row numbers starting from 1
    df['idx'] = range(1, len(df) + 1)
    
    # Set default values for potentially missing columns
    for col in ['timestamp', 'chatgpt_before', 'user', 'chatgpt_after']:
        if col not in df.columns:
            df[col] = ''
    
    # Specify the column order as shown in the screenshot - removing SRL_action
    columns = [
        'case_id',
        'week',
        'idx',
        'timestamp',
        'chatgpt_before',
        'user',
        'chatgpt_after'
    ]
    
    # Select only the columns in the specified order
    # Make sure all required columns exist
    final_columns = []
    for col in columns:
        if col in df.columns:
            final_columns.append(col)
        else:
            print(f"Warning: Column '{col}' not found in data. Skipping this column.")
    
    df = df[final_columns]
    
    # A completely different approach - split cells with very long text
    # This will help Excel calculate row heights correctly
    for col in ['chatgpt_before', 'user', 'chatgpt_after']:
        if col in df.columns:
            # Add explicit newlines every ~100 characters to help with wrapping
            df[col] = df[col].apply(lambda x: '\n'.join([x[i:i+100] for i in range(0, len(str(x)), 100)]) if isinstance(x, str) else x)
    
    try:
        # Use openpyxl for complete control
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        
        # Write the data to the worksheet
        rows = dataframe_to_rows(df, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Set reasonable column widths
        ws.column_dimensions['A'].width = 15  # case_id
        ws.column_dimensions['B'].width = 8   # week
        ws.column_dimensions['C'].width = 8   # idx
        ws.column_dimensions['D'].width = 15  # timestamp
        ws.column_dimensions['E'].width = 80  # chatgpt_before
        ws.column_dimensions['F'].width = 80  # user
        ws.column_dimensions['G'].width = 80  # chatgpt_after
        
        # Set Arial 10 font for all cells
        arial_10 = Font(name='Arial', size=10)
        header_font = Font(name='Arial', size=10, bold=True)
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting to all data cells
        for row in range(2, ws.max_row + 1):
            # Set row height to None to activate auto-sizing
            ws.row_dimensions[row].height = None
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = arial_10
                
                # Apply text wrapping to help with row height calculation
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Save the workbook
        wb.save(output_file)
        
        print(f"Data saved to {output_file}")
        print(f"Total rows saved: {len(df)}")
        print(f"Row heights are set to auto-adjust based on content")
        
    except Exception as e:
        print(f"Error when creating Excel file: {e}")
        print("Falling back to basic pandas Excel export")
        
        # Use pandas with the engine that supports auto row height better
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                # Get access to the workbook
                workbook = writer.book
                worksheet = workbook.active
                
                # Apply styling
                for col in range(1, 8):
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 8
                worksheet.column_dimensions['C'].width = 8
                worksheet.column_dimensions['D'].width = 15
                worksheet.column_dimensions['E'].width = 80
                worksheet.column_dimensions['F'].width = 80
                worksheet.column_dimensions['G'].width = 80
                
            print(f"Data saved to {output_file} using fallback method")
        except Exception as e2:
            print(f"Error with fallback method: {e2}")
            # Last resort
            df.to_excel(output_file, index=False)
            print(f"Data saved with basic formatting only")


def main():
    # Base directory where task directories are located
    base_dir = "."  # Change this to the actual path
    
    print("Starting to process files...")
    
    # Process all files
    all_data = process_all_files(base_dir)
    
    print(f"Found {len(all_data)} valid entries from the files")
    
    # Adjust timestamps based on week and idx
    all_data = adjust_timestamps(all_data)
    
    print("Timestamps adjusted")
    
    # Save to Excel - using a new Excel filename
    output_file = "processed_users_data.xlsx"
    save_to_excel(all_data, output_file)
    
    print(f"Processing complete. All data saved to {output_file}")


if __name__ == "__main__":
    main()